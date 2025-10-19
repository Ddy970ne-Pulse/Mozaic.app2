"""
📊 ADAPTATEUR TABLEAU ANALYSE ABSENCES
Transforme TableauAnalyseAbsences en format double-bloc :
- ABSENCES PROGRAMMÉES (gauche) : CA, FO, RTT, etc.
- ABSENTÉISME (droit) : AM, NAUT, etc.

Usage:
    python adapt_absences_tableau.py --source data/absences.xlsx --mapping config/mapping.csv --output out/Analyse_Final.xlsx
"""

import pandas as pd
import numpy as np
import json
import argparse
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 📋 CONFIGURATION DES TYPES D'ABSENCE
TYPES_PROGRAMMEES = [
    'CA',   # Congés Annuels
    'CT',   # Congés Trimestriels
    'RTT',  # RTT
    'REC',  # Récupération
    'FO',   # Formation
    'MAT',  # Congé maternité
    'PAT',  # Congé paternité
    'FAM',  # Événement familial
    'RMED', # Rendez-vous médical
    'EMAL', # Enfants malades
    'CEX',  # Congé exceptionnel
    'RH',   # Repos Hebdomadaire
    'RHD',  # Repos Dominical
    'STG',  # Stage
    'TEL',  # Télétravail
    'DEL'   # Délégation
]

TYPES_ABSENTEISME = [
    'AM',   # Arrêt maladie
    'AT',   # Accident du travail
    'MPRO', # Maladie Professionnelle
    'NAUT', # Absence non autorisée
    'AUT',  # Absence autorisée
    'CSS'   # Congés Sans Solde
]

# Mapping par défaut si fichier mapping non fourni
DEFAULT_MAPPING = {
    'Congés Annuels': 'CA',
    'Congés annuels': 'CA',
    'Congés Payés': 'CA',
    'Congés payés': 'CA',
    'CA - Congés Annuels': 'CA',
    'Congés Trimestriels': 'CT',
    'Congés trimestriels': 'CT',
    'RTT': 'RTT',
    'RTT (Réduction Temps Travail)': 'RTT',
    'Récupération': 'REC',
    'Formation': 'FO',
    'Congé formation': 'FO',
    'Congé maternité': 'MAT',
    'Congé paternité': 'PAT',
    'Évènement familial': 'FAM',
    'Événement familial': 'FAM',
    'Rendez-vous médical': 'RMED',
    'Enfants malades': 'EMAL',
    'Congé exceptionnel': 'CEX',
    'Repos Hebdomadaire': 'RH',
    'Repos Dominical': 'RHD',
    'Stage': 'STG',
    'Télétravail': 'TEL',
    'Délégation': 'DEL',
    'Arrêt maladie': 'AM',
    'Accident du travail/Trajet': 'AT',
    'Accident du travail': 'AT',
    'Maladie Professionnelle': 'MPRO',
    'Absence non autorisée': 'NAUT',
    'Absence autorisée': 'AUT',
    'Congés Sans Solde': 'CSS'
}

class AbsenceTableauAdapter:
    def __init__(self, source_path, mapping_path=None, output_path=None, counting_method='Jours Ouvrés'):
        self.source_path = Path(source_path)
        self.mapping_path = Path(mapping_path) if mapping_path else None
        self.output_path = Path(output_path) if output_path else Path('./out/Analyse_Absences_Final.xlsx')
        self.counting_method = counting_method
        
        # Logs
        self.log = {
            'timestamp': datetime.now().isoformat(),
            'source_file': str(self.source_path),
            'output_file': str(self.output_path),
            'rows_processed': 0,
            'unmapped_types': [],
            'employees_processed': 0,
            'durations_calculated': 0,
            'warnings': [],
            'errors': []
        }
        
        # Data
        self.df_source = None
        self.mapping = None
        self.df_final = None
    
    def load_source_data(self):
        """Charge le fichier source (CSV ou XLSX)"""
        print(f"📂 Chargement de {self.source_path}...")
        
        try:
            if self.source_path.suffix.lower() == '.csv':
                self.df_source = pd.read_csv(self.source_path)
            elif self.source_path.suffix.lower() in ['.xlsx', '.xls']:
                self.df_source = pd.read_excel(self.source_path)
            else:
                raise ValueError(f"Format non supporté: {self.source_path.suffix}")
            
            print(f"✅ {len(self.df_source)} lignes chargées")
            self.log['rows_processed'] = len(self.df_source)
            
            # Vérifier colonnes essentielles
            required_cols = ['EmployeNom', 'TypeAbsence']
            missing_cols = [col for col in required_cols if col not in self.df_source.columns]
            
            if missing_cols:
                # Essayer noms alternatifs
                col_alternatives = {
                    'EmployeNom': ['employee_name', 'nom', 'name', 'Nom', 'Employee'],
                    'TypeAbsence': ['type', 'motif', 'motif_absence', 'Type', 'Motif'],
                    'Duree': ['duree', 'jours', 'days', 'duration', 'jours_absence']
                }
                
                for required, alternatives in col_alternatives.items():
                    if required not in self.df_source.columns:
                        for alt in alternatives:
                            if alt in self.df_source.columns:
                                self.df_source.rename(columns={alt: required}, inplace=True)
                                print(f"   ℹ️  Colonne '{alt}' renommée en '{required}'")
                                break
                
                # Re-vérifier
                missing_cols = [col for col in required_cols if col not in self.df_source.columns]
                if missing_cols:
                    raise ValueError(f"Colonnes manquantes: {missing_cols}")
            
            return self.df_source
            
        except Exception as e:
            self.log['errors'].append(f"Erreur chargement source: {str(e)}")
            raise
    
    def load_mapping(self):
        """Charge le fichier de mapping ou utilise le mapping par défaut"""
        if self.mapping_path and self.mapping_path.exists():
            print(f"📋 Chargement du mapping depuis {self.mapping_path}...")
            
            try:
                if self.mapping_path.suffix.lower() == '.csv':
                    df_mapping = pd.read_csv(self.mapping_path)
                elif self.mapping_path.suffix.lower() == '.json':
                    with open(self.mapping_path, 'r', encoding='utf-8') as f:
                        self.mapping = json.load(f)
                    print(f"✅ {len(self.mapping)} mappings chargés depuis JSON")
                    return self.mapping
                else:
                    raise ValueError(f"Format mapping non supporté: {self.mapping_path.suffix}")
                
                # Convertir DataFrame en dict
                if 'Intitule' in df_mapping.columns and 'Code' in df_mapping.columns:
                    self.mapping = dict(zip(df_mapping['Intitule'], df_mapping['Code']))
                else:
                    self.mapping = dict(zip(df_mapping.iloc[:, 0], df_mapping.iloc[:, 1]))
                
                print(f"✅ {len(self.mapping)} mappings chargés")
                return self.mapping
                
            except Exception as e:
                self.log['warnings'].append(f"Erreur chargement mapping: {str(e)}, utilisation mapping par défaut")
                print(f"⚠️  {self.log['warnings'][-1]}")
        
        # Utiliser mapping par défaut
        self.mapping = DEFAULT_MAPPING.copy()
        print(f"✅ Mapping par défaut utilisé ({len(self.mapping)} entrées)")
        return self.mapping
    
    def map_type_codes(self):
        """Mappe les types d'absence vers les codes officiels"""
        print("\n🔄 Mapping des types d'absence...")
        
        def map_type(type_str):
            if pd.isna(type_str):
                return 'INCONNU'
            
            type_str = str(type_str).strip()
            
            # Recherche exacte
            if type_str in self.mapping:
                return self.mapping[type_str]
            
            # Recherche insensible à la casse
            for intitule, code in self.mapping.items():
                if type_str.lower() == intitule.lower():
                    return code
            
            # Type non mappé
            if type_str not in self.log['unmapped_types']:
                self.log['unmapped_types'].append(type_str)
                print(f"   ⚠️  Type non mappé: '{type_str}'")
            
            return 'INCONNU'
        
        self.df_source['CodeTypeAbsence'] = self.df_source['TypeAbsence'].apply(map_type)
        
        mapped_count = (self.df_source['CodeTypeAbsence'] != 'INCONNU').sum()
        print(f"✅ {mapped_count}/{len(self.df_source)} types mappés avec succès")
        
        if self.log['unmapped_types']:
            print(f"⚠️  {len(self.log['unmapped_types'])} types non mappés")
        
        return self.df_source
    
    def classify_absences(self):
        """Classifie en PROGRAMMÉES vs ABSENTÉISME"""
        print("\n📊 Classification des absences...")
        
        def classify(row):
            code = row['CodeTypeAbsence']
            
            # Si StatutPlanif existe, le respecter
            if 'StatutPlanif' in row and pd.notna(row['StatutPlanif']):
                statut = str(row['StatutPlanif']).lower()
                if 'planif' in statut or 'valid' in statut or 'appro' in statut:
                    return 'programmee'
                else:
                    return 'absenteisme'
            
            # Classification par code
            if code in TYPES_PROGRAMMEES:
                return 'programmee'
            elif code in TYPES_ABSENTEISME:
                return 'absenteisme'
            else:
                # Par défaut : absentéisme pour INCONNU
                return 'absenteisme'
        
        self.df_source['Classification'] = self.df_source.apply(classify, axis=1)
        
        prog_count = (self.df_source['Classification'] == 'programmee').sum()
        absent_count = (self.df_source['Classification'] == 'absenteisme').sum()
        
        print(f"✅ Programmées: {prog_count}, Absentéisme: {absent_count}")
        
        return self.df_source
    
    def calculate_durations(self):
        """Calcule les durées si manquantes"""
        print("\n⏱️  Calcul des durées...")
        
        if 'Duree' not in self.df_source.columns:
            # Chercher colonnes alternatives
            for col in ['duree', 'jours', 'jours_absence', 'duration']:
                if col in self.df_source.columns:
                    self.df_source.rename(columns={col: 'Duree'}, inplace=True)
                    break
        
        if 'Duree' not in self.df_source.columns:
            self.df_source['Duree'] = 0.0
        
        # Convertir en numérique
        self.df_source['Duree'] = pd.to_numeric(self.df_source['Duree'], errors='coerce').fillna(0)
        
        # Si Duree = 0 et dates disponibles, calculer
        if 'DateDebut' in self.df_source.columns and 'DateFin' in self.df_source.columns:
            mask_missing = self.df_source['Duree'] == 0
            
            for idx in self.df_source[mask_missing].index:
                try:
                    date_debut = pd.to_datetime(self.df_source.loc[idx, 'DateDebut'])
                    date_fin = pd.to_datetime(self.df_source.loc[idx, 'DateFin'])
                    
                    if pd.notna(date_debut) and pd.notna(date_fin):
                        # Calcul simple : différence en jours + 1
                        days = (date_fin - date_debut).days + 1
                        
                        # TODO: Implémenter counting_method (Jours Ouvrés vs Calendaires)
                        # Pour l'instant, utilisation simple
                        
                        self.df_source.loc[idx, 'Duree'] = max(days, 1)
                        self.log['durations_calculated'] += 1
                        
                except Exception as e:
                    self.log['warnings'].append(f"Erreur calcul durée ligne {idx}: {str(e)}")
        
        if self.log['durations_calculated'] > 0:
            print(f"✅ {self.log['durations_calculated']} durées calculées")
        
        return self.df_source
    
    def create_pivot_tables(self):
        """Crée les deux pivots (programmées et absentéisme)"""
        print("\n📈 Création des pivots...")
        
        # Pivot PROGRAMMÉES
        df_prog = self.df_source[self.df_source['Classification'] == 'programmee']
        
        pivot_prog = df_prog.pivot_table(
            index='EmployeNom',
            columns='CodeTypeAbsence',
            values='Duree',
            aggfunc='sum',
            fill_value=0
        )
        
        # Ajouter colonnes manquantes avec 0
        for code in TYPES_PROGRAMMEES:
            if code not in pivot_prog.columns:
                pivot_prog[code] = 0
        
        # Trier colonnes dans l'ordre de TYPES_PROGRAMMEES
        cols_prog = [c for c in TYPES_PROGRAMMEES if c in pivot_prog.columns]
        pivot_prog = pivot_prog[cols_prog]
        
        # Ajouter Total
        pivot_prog['Total'] = pivot_prog.sum(axis=1)
        
        print(f"   ✅ Pivot programmées: {len(pivot_prog)} employés, {len(cols_prog)} types")
        
        # Pivot ABSENTÉISME
        df_absent = self.df_source[self.df_source['Classification'] == 'absenteisme']
        
        pivot_absent = df_absent.pivot_table(
            index='EmployeNom',
            columns='CodeTypeAbsence',
            values='Duree',
            aggfunc='sum',
            fill_value=0
        )
        
        # Ajouter colonnes manquantes avec 0
        for code in TYPES_ABSENTEISME:
            if code not in pivot_absent.columns:
                pivot_absent[code] = 0
        
        # Gérer INCONNU s'il existe
        if 'INCONNU' in pivot_absent.columns:
            cols_absent = [c for c in TYPES_ABSENTEISME if c in pivot_absent.columns] + ['INCONNU']
        else:
            cols_absent = [c for c in TYPES_ABSENTEISME if c in pivot_absent.columns]
        
        pivot_absent = pivot_absent[cols_absent]
        
        # Ajouter TOTAL
        pivot_absent['TOTAL'] = pivot_absent.sum(axis=1)
        
        print(f"   ✅ Pivot absentéisme: {len(pivot_absent)} employés, {len(cols_absent)} types")
        
        # Unifier les index (tous les employés dans les deux)
        all_employees = sorted(set(pivot_prog.index) | set(pivot_absent.index))
        
        pivot_prog = pivot_prog.reindex(all_employees, fill_value=0)
        pivot_absent = pivot_absent.reindex(all_employees, fill_value=0)
        
        self.log['employees_processed'] = len(all_employees)
        print(f"✅ {len(all_employees)} employés uniques traités")
        
        return pivot_prog, pivot_absent
    
    def assemble_final_table(self, pivot_prog, pivot_absent):
        """Assemble les deux pivots côte à côte"""
        print("\n🔗 Assemblage du tableau final...")
        
        # Reset index pour avoir EmployeNom comme colonne
        pivot_prog = pivot_prog.reset_index()
        pivot_absent = pivot_absent.reset_index()
        
        # Renommer colonnes absentéisme pour éviter conflits
        pivot_absent = pivot_absent.rename(columns={'EmployeNom': 'EmployeNom_temp'})
        
        # Joindre
        self.df_final = pd.merge(
            pivot_prog,
            pivot_absent.drop('EmployeNom_temp', axis=1),
            left_index=True,
            right_index=True,
            how='outer'
        )
        
        # Remplacer NaN par 0
        self.df_final = self.df_final.fillna(0)
        
        # Arrondir à 1 décimale
        numeric_cols = self.df_final.select_dtypes(include=[np.number]).columns
        self.df_final[numeric_cols] = self.df_final[numeric_cols].round(1)
        
        print(f"✅ Tableau final: {len(self.df_final)} lignes, {len(self.df_final.columns)} colonnes")
        
        return self.df_final
    
    def export_to_excel(self):
        """Exporte vers Excel avec formatage"""
        print(f"\n💾 Export vers {self.output_path}...")
        
        # Créer répertoire si nécessaire
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Créer workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Analyse"
        
        # Styles
        header_prog_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
        header_absent_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Vert clair
        total_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Or
        
        header_font = Font(bold=True, size=11)
        total_font = Font(bold=True, size=10)
        
        center_align = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Écrire en-têtes
        headers = list(self.df_final.columns)
        
        # Trouver index de séparation (après 'Total' des programmées)
        try:
            sep_idx = headers.index('Total') + 1
        except:
            sep_idx = len([c for c in headers if c in TYPES_PROGRAMMEES]) + 1
        
        # Ligne 1 : Titres de blocs
        ws['A1'] = 'ABSENCES PROGRAMMÉES'
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sep_idx)
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].fill = header_prog_fill
        ws['A1'].alignment = center_align
        
        ws.cell(row=1, column=sep_idx+1, value='ABSENTÉISME')
        ws.merge_cells(start_row=1, start_column=sep_idx+1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=sep_idx+1).font = Font(bold=True, size=12)
        ws.cell(row=1, column=sep_idx+1).fill = header_absent_fill
        ws.cell(row=1, column=sep_idx+1).alignment = center_align
        
        # Ligne 2 : En-têtes de colonnes
        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
            if col_idx <= sep_idx:
                cell.fill = header_prog_fill
            else:
                cell.fill = header_absent_fill
        
        # Données
        for row_idx, row_data in enumerate(self.df_final.itertuples(index=False), start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # Formatage des totaux
                if headers[col_idx-1] in ['Total', 'TOTAL']:
                    cell.font = total_font
                    cell.fill = total_fill
                
                # Centrage sauf EmployeNom
                if col_idx > 1:
                    cell.alignment = center_align
        
        # Ajuster largeurs
        ws.column_dimensions['A'].width = 25  # EmployeNom
        for col_idx in range(2, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = 8
        
        # Sauvegarder
        wb.save(self.output_path)
        
        print(f"✅ Fichier Excel créé: {self.output_path}")
        
        return self.output_path
    
    def save_log(self):
        """Sauvegarde le log JSON"""
        log_path = Path('./out/absences_adapt_log.json')
        log_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(log_path, 'w', encoding='utf-8') as f:
            json.dump(self.log, f, indent=2, ensure_ascii=False)
        
        print(f"\n📋 Log sauvegardé: {log_path}")
        
        return log_path
    
    def run(self):
        """Exécution complète du pipeline"""
        print("=" * 80)
        print("📊 ADAPTATEUR TABLEAU ANALYSE ABSENCES")
        print("=" * 80)
        
        try:
            # 1. Charger données
            self.load_source_data()
            
            # 2. Charger mapping
            self.load_mapping()
            
            # 3. Mapper types
            self.map_type_codes()
            
            # 4. Classifier
            self.classify_absences()
            
            # 5. Calculer durées
            self.calculate_durations()
            
            # 6. Créer pivots
            pivot_prog, pivot_absent = self.create_pivot_tables()
            
            # 7. Assembler
            self.assemble_final_table(pivot_prog, pivot_absent)
            
            # 8. Exporter Excel
            self.export_to_excel()
            
            # 9. Sauvegarder log
            self.save_log()
            
            print("\n" + "=" * 80)
            print("✅ TRAITEMENT TERMINÉ AVEC SUCCÈS")
            print("=" * 80)
            print(f"\n📊 RÉSUMÉ:")
            print(f"   • Lignes traitées: {self.log['rows_processed']}")
            print(f"   • Employés uniques: {self.log['employees_processed']}")
            print(f"   • Types non mappés: {len(self.log['unmapped_types'])}")
            print(f"   • Durées calculées: {self.log['durations_calculated']}")
            print(f"   • Fichier final: {self.output_path}")
            
            if self.log['unmapped_types']:
                print(f"\n⚠️  TYPES NON MAPPÉS:")
                for t in self.log['unmapped_types'][:10]:
                    print(f"   - {t}")
                if len(self.log['unmapped_types']) > 10:
                    print(f"   ... et {len(self.log['unmapped_types']) - 10} autres")
            
            return True
            
        except Exception as e:
            print(f"\n❌ ERREUR: {str(e)}")
            self.log['errors'].append(str(e))
            self.save_log()
            raise


def main():
    parser = argparse.ArgumentParser(description='Adaptateur Tableau Analyse Absences')
    parser.add_argument('--source', required=True, help='Chemin du fichier source (CSV/XLSX)')
    parser.add_argument('--mapping', help='Chemin du fichier de mapping (CSV/JSON, optionnel)')
    parser.add_argument('--output', default='./out/Analyse_Absences_Final.xlsx', help='Chemin du fichier Excel de sortie')
    parser.add_argument('--counting-method', default='Jours Ouvrés', choices=['Jours Ouvrés', 'Jours Calendaires'], help='Méthode de calcul des durées')
    
    args = parser.parse_args()
    
    adapter = AbsenceTableauAdapter(
        source_path=args.source,
        mapping_path=args.mapping,
        output_path=args.output,
        counting_method=args.counting_method
    )
    
    adapter.run()


if __name__ == "__main__":
    main()
